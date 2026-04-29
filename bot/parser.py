"""Excel → list[CarDTO] parser.

Reads the first sheet. Columns are matched by header name (RU/EN aliases),
so source files can have any column order and extra columns.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

# Supported header aliases (normalized: lowercased, stripped, punctuation removed).
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "brand": ("brand", "make", "марка", "марка автомобиля", "бренд"),
    "model": ("model", "модель"),
    "price": (
        "price",
        "цена",
        "цена продажи",
        "цена продажи руб",
        "стоимость",
        "стоимость продажи",
    ),
    "year": ("year", "год", "год выпуска"),
    "mileage": ("mileage", "пробег", "пробег км"),
    "fuel": ("fuel", "топливо", "тип топлива", "двигатель"),
    "transmission": ("transmission", "коробка", "кпп", "трансмиссия", "тип кпп"),
    "vin": ("vin", "вин", "номер vin"),
    "location": (
        "location",
        "город",
        "местоположение",
        "месторасположение",
        "регион",
    ),
}

REQUIRED_FIELDS = ("brand", "model", "price", "year", "mileage", "vin")

# Suffixes / units we strip before parsing numbers.
_NUMERIC_NOISE_RE = re.compile(
    r"(руб(лей|ля|\.)?|р\.?|г(ода|\.)?|км\.?|тыс\.?|\s|,)",
    flags=re.IGNORECASE,
)


@dataclass
class CarDTO:
    brand: str
    model: str
    price: int
    year: int
    mileage: int
    vin: str
    fuel: str | None = None
    transmission: str | None = None
    location: str | None = None

    def to_payload(self) -> dict[str, Any]:
        payload: dict[str, Any] = {
            "brand": self.brand,
            "model": self.model,
            "price": self.price,
            "year": self.year,
            "mileage": self.mileage,
            "vin": self.vin,
        }
        if self.fuel:
            payload["fuel"] = self.fuel
        if self.transmission:
            payload["transmission"] = self.transmission
        if self.location:
            payload["location"] = self.location
        return payload


@dataclass
class ParseResult:
    cars: list[CarDTO] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def _normalize_header(cell: Any) -> str:
    if cell is None:
        return ""
    text = str(cell)
    # Replace non-breaking / zero-width / weird whitespace with regular space.
    text = text.replace("\u00a0", " ").replace("\u200b", "").replace("\ufeff", "")
    text = text.strip().lower().strip('"\'«»').strip()
    # Drop any punctuation (commas, dots, slashes, hashes etc), collapse whitespace.
    text = re.sub(r"[.,;:!?()\[\]/\\№#*]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _detect_header(row: tuple[Any, ...]) -> dict[str, int] | None:
    """Return a mapping {field_name: column_index}, or None if no header found."""
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row):
        norm = _normalize_header(cell)
        if not norm:
            continue
        for field_name, aliases in COLUMN_ALIASES.items():
            if field_name in mapping:
                continue
            # Match if normalized header equals alias OR starts with alias+space
            # (e.g. "год выпуска" alias "год" → also matches, but full alias "год выпуска" wins via ordering).
            if norm in aliases:
                mapping[field_name] = idx
                break
    # A row is a header only if we matched at least brand + vin or brand + price.
    if "brand" in mapping and ("vin" in mapping or "price" in mapping):
        return mapping
    return None


# How many top rows to scan when looking for the header row.
_HEADER_SCAN_ROWS = 10


def _to_int(value: Any) -> int:
    if value is None or value == "":
        raise ValueError("пусто")
    if isinstance(value, bool):  # bool is a subclass of int, treat specially
        raise ValueError(f"не число: {value!r}")
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        cleaned = _NUMERIC_NOISE_RE.sub("", value).replace("\u00a0", "")
        if not cleaned:
            raise ValueError(f"пусто (было {value!r})")
        try:
            return int(float(cleaned))
        except ValueError as exc:
            raise ValueError(f"не число: {value!r}") from exc
    raise ValueError(f"не число: {value!r}")


def _to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _col_letter(idx: int) -> str:
    """Convert 0-based column index to Excel letter (0 → 'A')."""
    result = ""
    n = idx
    while True:
        result = chr(ord("A") + n % 26) + result
        n = n // 26 - 1
        if n < 0:
            break
    return result


def _positional_mapping() -> dict[str, int]:
    """Legacy fallback: A..I = brand, model, price, year, mileage, fuel, transmission, vin, location."""
    return {
        "brand": 0,
        "model": 1,
        "price": 2,
        "year": 3,
        "mileage": 4,
        "fuel": 5,
        "transmission": 6,
        "vin": 7,
        "location": 8,
    }


def _get(row: tuple[Any, ...], mapping: dict[str, int], field_name: str) -> Any:
    idx = mapping.get(field_name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def parse_excel(data: bytes) -> ParseResult:
    """Parse the first sheet of an xlsx file into CarDTO objects.

    Scans the first rows looking for a header row (matched by column names,
    see COLUMN_ALIASES). If no header is found, falls back to the legacy
    positional layout (A..I) only when the first row looks numeric.
    """
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb.worksheets[0]

    result = ParseResult()
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return result

    # Scan top rows for a header match.
    header_row_idx: int | None = None
    mapping: dict[str, int] | None = None
    for i, row in enumerate(all_rows[:_HEADER_SCAN_ROWS]):
        candidate = _detect_header(row)
        if candidate is not None:
            mapping = candidate
            header_row_idx = i
            break

    if mapping is None:
        # No header anywhere in the top rows. Fall back to positional ONLY if
        # the first non-empty row looks like numeric data; otherwise report a
        # clear error with what we saw so the user can adjust.
        first_nonempty = next(
            (r for r in all_rows if r and any(c is not None and c != "" for c in r)),
            None,
        )
        if first_nonempty is None:
            return result

        # Heuristic: treat as positional data if col C (price) is numeric.
        col_c = first_nonempty[2] if len(first_nonempty) > 2 else None
        looks_numeric = isinstance(col_c, (int, float)) or (
            isinstance(col_c, str) and any(ch.isdigit() for ch in col_c)
        )
        if not looks_numeric:
            preview = [
                _normalize_header(c) for c in all_rows[0][:12] if c not in (None, "")
            ]
            result.errors.append(
                "Не удалось найти строку заголовков. Ожидаю как минимум "
                "колонки 'Марка' и 'VIN' (или 'Цена продажи') в первых "
                f"{_HEADER_SCAN_ROWS} строках. Увидел в первой строке: "
                + (", ".join(repr(p) for p in preview) if preview else "пусто")
            )
            return result

        mapping = _positional_mapping()
        data_rows: list[tuple[Any, ...]] = all_rows
        start_row = 1
    else:
        missing = [f for f in REQUIRED_FIELDS if f not in mapping]
        if missing:
            found = ", ".join(f"{k}={_col_letter(v)}" for k, v in mapping.items())
            result.errors.append(
                f"В заголовке не хватает обязательных колонок: {', '.join(missing)}. "
                f"Нашёл: {found}."
            )
            return result
        assert header_row_idx is not None
        data_rows = all_rows[header_row_idx + 1 :]
        start_row = header_row_idx + 2

    for offset, raw in enumerate(data_rows):
        row_idx = start_row + offset
        if raw is None or all(cell is None or cell == "" for cell in raw):
            continue
        try:
            brand = _to_str(_get(raw, mapping, "brand"))
            model = _to_str(_get(raw, mapping, "model"))
            vin = _to_str(_get(raw, mapping, "vin"))
            if not brand or not model or not vin:
                raise ValueError("пустые brand/model/vin")
            car = CarDTO(
                brand=brand,
                model=model,
                price=_to_int(_get(raw, mapping, "price")),
                year=_to_int(_get(raw, mapping, "year")),
                mileage=_to_int(_get(raw, mapping, "mileage")),
                vin=vin,
                fuel=_to_str(_get(raw, mapping, "fuel")) or None,
                transmission=_to_str(_get(raw, mapping, "transmission")) or None,
                location=_to_str(_get(raw, mapping, "location")) or None,
            )
        except (ValueError, TypeError) as exc:
            result.errors.append(f"строка {row_idx}: {exc}")
            continue
        result.cars.append(car)

    return result
